import json
import os
import openpyxl
from datetime import datetime, date

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SRC = "/sessions/wizardly-compassionate-mendel/mnt/Vacanze/Vacanza 2026.xlsx"

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb["Itinerario 2026"]

# Coordinates (approximate, WGS84) for places mentioned in the itinerary.
# NOTE: these are place-level markers for the overview map. Real GPX tracks
# will replace the straight connecting lines once provided by the user.
COORDS = {
    "Alessandria": [44.9133, 8.6141],
    "Milano Rogoredo": [45.4342, 9.2286],
    "Milano Lambrate": [45.4874, 9.2373],
    "Verona Porta Nuova": [45.4299, 10.9840],
    "Venezia Mestre": [45.4832, 12.2357],
    "Cervignano-Aquileia-Grado": [45.8258, 13.3313],
    "Radetzky Rooms": [45.7686, 13.3711],
    "Basilica Patriarcale di Santa Maria Assunta": [45.7669, 13.3714],
    "Pontile ASNAT": [45.7040, 13.3700],
    "Marano Lagunare": [45.7342, 13.1758],
    "Jo Hotel": [45.7346, 13.1755],
    "Lignano Sabbiadoro": [45.6773, 13.1345],
    "Hotel Marina Uno": [45.6790, 13.1310],
    "Bibione (terminal traghetti)": [45.6156, 13.0428],
    "ValleVecchia (molo traghetti)": [45.6667, 12.9833],
    "ValleVecchia (imbarco traghetto Caorle )": [45.6520, 12.9200],
    "Caorle": [45.5985, 12.8892],
    "Appartamento (Caorle)": [45.5985, 12.8892],
    "Casa del Marinaio": [45.5940, 12.8710],
    "Imbarco Porto Santa Margherita": [45.5860, 12.8520],
    "Porto Santa Margherita": [45.5840, 12.8500],
    "Cortellazzo": [45.5324, 12.6892],
    "Punta Tre Porti": [45.4599, 12.5423],
    "Burano": [45.4854, 12.4166],
    "Venezia": [45.4408, 12.3155],
    "Bologna Centrale": [44.5075, 11.3426],
}

def coord_lookup(name):
    if not name:
        return None
    name = name.strip()
    if name in COORDS:
        return COORDS[name]
    for k, v in COORDS.items():
        if k.lower() in name.lower() or name.lower() in k.lower():
            return v
    return None

def to_hhmm(v):
    if v is None:
        return None
    if hasattr(v, "strftime"):
        return v.strftime("%H:%M")
    return str(v)

def compute_duration(h, i):
    # Fallback: if the sheet has both a partenza and an arrivo time but no
    # explicit durata, compute it ourselves (same-day, no overnight legs
    # in this itinerary).
    if not (hasattr(h, "hour") and hasattr(i, "hour")):
        return None
    start_min = h.hour * 60 + h.minute
    end_min = i.hour * 60 + i.minute
    diff = end_min - start_min
    if diff < 0:
        return None
    return f"{diff // 60:02d}:{diff % 60:02d}"

MESI_IT = [
    "gennaio", "febbraio", "marzo", "aprile", "maggio", "giugno",
    "luglio", "agosto", "settembre", "ottobre", "novembre", "dicembre",
]

def date_label_it(c):
    # e.g. "4 agosto" — no leading zero, Italian month name, no year
    # (the trip only spans 2026, so the year is omitted as redundant).
    return f"{c.day} {MESI_IT[c.month - 1]}"

# Link di prenotazione e servizi (cucina/lavatrice/colazione) per ciascun
# alloggio, forniti manualmente dall'utente: non sono nel foglio Excel.
ACCOMMODATION_INFO = {
    "Radetzky Rooms": {
        "link": "https://www.booking.com/Share-m15lyw",
        "servizi": "casa con cucina e lavatrice",
    },
    "Jo Hotel": {
        "link": "https://www.booking.com/Share-6J0jbgU",
        "servizi": "prima colazione inclusa",
    },
    "Hotel Marina Uno": {
        "link": "https://www.booking.com/Share-78YSCe",
        "servizi": "prima colazione inclusa",
    },
    "Hotel Capinera": {
        "link": "https://www.booking.com/Share-NTGp86",
        "servizi": "casa con cucina e lavatrice",
    },
    "Iesolo: Casa": {
        "link": "https://www.airbnb.it/rooms/1701657293748998258",
        "servizi": "casa con cucina e lavatrice",
    },
    "Welcome Home": {
        "link": "https://www.booking.com/Share-Fdm8Wa",
        "servizi": "casa con cucina e lavatrice",
    },
    "Dimora Veneziana Rooms": {
        "link": "https://www.booking.com/Share-FibbB4",
        "servizi": "casa con cucina e lavatrice",
    },
}

def accommodation_extra(nome):
    if not nome:
        return {"link": None, "servizi": None}
    nome = nome.strip()
    if nome in ACCOMMODATION_INFO:
        return ACCOMMODATION_INFO[nome]
    for k, v in ACCOMMODATION_INFO.items():
        if k.lower() in nome.lower() or nome.lower() in k.lower():
            return v
    return {"link": None, "servizi": None}

days = []
current_day = None

MAX_ROW = ws.max_row  # dinamico: si adatta se il foglio cresce o si accorcia

for r in range(3, MAX_ROW + 1):
    b = ws[f"B{r}"].value
    c = ws[f"C{r}"].value
    d = ws[f"D{r}"].value

    if c is not None and isinstance(c, datetime):
        if current_day:
            days.append(current_day)
        current_day = {
            "date": c.strftime("%Y-%m-%d"),
            "date_label": date_label_it(c),
            "weekday": b,
            "tappa": d,
            "legs": [],
            "accommodations": [],
            "expenses": [],
        }

    if current_day is None:
        continue

    e = ws[f"E{r}"].value
    f = ws[f"F{r}"].value
    g = ws[f"G{r}"].value
    h = ws[f"H{r}"].value
    i = ws[f"I{r}"].value
    j = ws[f"J{r}"].value
    k = ws[f"K{r}"].value
    l = ws[f"L{r}"].value
    m = ws[f"M{r}"].value
    n = ws[f"N{r}"].value
    o = ws[f"O{r}"].value
    p = ws[f"P{r}"].value

    if e:
        origin, dest = None, None
        if "→" in e:
            parts = [x.strip() for x in e.split("→")]
            origin = parts[0] if parts[0] else None
            dest = parts[1] if len(parts) > 1 and parts[1] else None
        leg = {
            "row": r,
            "percorso": e.strip(),
            "origin": origin,
            "destination": dest,
            "origin_coords": coord_lookup(origin) if origin else None,
            "dest_coords": coord_lookup(dest) if dest else None,
            "mezzo": f,
            "km": round(g, 2) if isinstance(g, (int, float)) else g,
            "partenza": to_hhmm(h),
            "arrivo": to_hhmm(i),
            "durata": to_hhmm(j) or compute_duration(h, i),
        }
        current_day["legs"].append(leg)

    if k:
        nome_acc = k.strip() if isinstance(k, str) else k
        extra = accommodation_extra(nome_acc)
        current_day["accommodations"].append({
            "row": r,
            "nome": nome_acc,
            "tipo": l,
            "categoria": m,
            "quantita": n,
            "importo": o,
            "note": p,
            "link": extra["link"],
            "servizi": extra["servizi"],
        })
    elif l and m and o is not None and not k:
        current_day["expenses"].append({
            "row": r,
            "voce": l,
            "categoria": m.strip().lower() if isinstance(m, str) else m,
            "quantita": n,
            "importo": o,
            "note": p,
        })

if current_day:
    days.append(current_day)

ws2 = wb["Costi preparatori 2026"]
prep_costs = []
for r in range(3, 11):
    b = ws2[f"B{r}"].value
    c = ws2[f"C{r}"].value
    d = ws2[f"D{r}"].value
    e = ws2[f"E{r}"].value
    if b:
        prep_costs.append({
            "voce": b,
            "categoria": c,
            "quantita": d,
            "importo": e if e is not None else 0,
        })

prep_total = ws2["E11"].value

ws3 = wb["Totali 2026"]
totals = {
    "costi_preparatori": ws3["D2"].value,
    "itinerario": ws3["D3"].value,
    "totale": ws3["D5"].value,
}

trip = {
    "meta": {
        "titolo": "Vacanza in bici 2026 — Alessandria, Aquileia, Laguna, Venezia",
        "generato_il": date.today().isoformat(),
    },
    "days": days,
    "costi_preparatori": prep_costs,
    "costi_preparatori_totale": prep_total,
    "totali": totals,
}

out_json = os.path.join(BASE_DIR, "data", "trip.json")
out_js = os.path.join(BASE_DIR, "js", "data.js")

os.makedirs(os.path.dirname(out_json), exist_ok=True)
with open(out_json, "w", encoding="utf-8") as f:
    json.dump(trip, f, ensure_ascii=False, indent=2)

with open(out_js, "w", encoding="utf-8") as f:
    f.write("window.TRIP_DATA = ")
    json.dump(trip, f, ensure_ascii=False, indent=2)
    f.write(";\n")

print("Days:", len(days))
print("Totals:", totals)
print("Wrote:", out_json)
print("Wrote:", out_js)
